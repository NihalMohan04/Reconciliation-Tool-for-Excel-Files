from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RED_BOLD_FONT = Font(color="00FF0000", bold=True)
HEADER_BOLD_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center")
ECC_FILL = PatternFill(fill_type="solid", start_color="00E0FFFF", end_color="00E0FFFF")
S4_FILL = PatternFill(fill_type="solid", start_color="00FFF2CC", end_color="00FFF2CC")


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sort_key(row_values: tuple[object, ...]) -> tuple[str, str, str]:
    first = _normalize_cell_value(row_values[0] if len(row_values) > 0 else "")
    second = _normalize_cell_value(row_values[1] if len(row_values) > 1 else "")
    third = _normalize_cell_value(row_values[2] if len(row_values) > 2 else "")
    return (first, second, third)


def _read_first_sheet(file_path: Path) -> tuple[list[str], list[tuple[object, ...]]]:
    workbook = load_workbook(file_path, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows: list[tuple[object, ...]] = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return [], []

        max_columns = max(len(row) for row in rows)
        padded_rows = [row + (None,) * (max_columns - len(row)) for row in rows]

        raw_header = padded_rows[0]
        headers = [
            _normalize_cell_value(value) if _normalize_cell_value(value) else f"Column{index + 1}"
            for index, value in enumerate(raw_header)
        ]
        data_rows = padded_rows[1:]
        return headers, data_rows
    finally:
        workbook.close()


def _ensure_file_exists(file_path: Path) -> None:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")


def _pad_row(row: tuple[object, ...], width: int) -> tuple[object, ...]:
    if len(row) >= width:
        return row
    return row + (None,) * (width - len(row))


def _normalized_row(row: Iterable[object]) -> tuple[str, ...]:
    return tuple(_normalize_cell_value(value) for value in row)


def _auto_size_columns(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        max_length = 0
        for row_index in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if isinstance(cell_value, str) and cell_value.startswith("="):
                cell_length = len("Not Matched")
            else:
                cell_length = len(str(cell_value)) if cell_value is not None else 0
            if cell_length > max_length:
                max_length = cell_length

        adjusted_width = max(12, min(max_length + 2, 60))
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def _style_output_worksheet(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        worksheet.cell(row=1, column=column_index).font = HEADER_BOLD_FONT

    for row_index in range(1, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=1).alignment = CENTER_ALIGNMENT

    for column_index in range(2, worksheet.max_column + 1):
        fill = ECC_FILL if (column_index - 2) % 2 == 0 else S4_FILL
        for row_index in range(1, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).fill = fill


def reconcile_workbooks(
    source_file: str,
    target_file: str,
    output_file: str,
) -> dict:
    source_path = Path(source_file)
    target_path = Path(target_file)
    output_path = Path(output_file)

    _ensure_file_exists(source_path)
    _ensure_file_exists(target_path)

    source_headers, source_rows = _read_first_sheet(source_path)
    target_headers, target_rows = _read_first_sheet(target_path)

    source_rows_sorted = sorted(source_rows, key=_sort_key)
    target_rows_sorted = sorted(target_rows, key=_sort_key)

    source_width = len(source_headers)
    target_width = len(target_headers)
    compare_width = max(source_width, target_width)
    pair_count = max(len(source_rows_sorted), len(target_rows_sorted))

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "Reconciliation"

    output_header = ["Match Status"]
    for column_offset in range(compare_width):
        source_header = (
            source_headers[column_offset]
            if column_offset < source_width
            else f"Column{column_offset + 1}"
        )
        target_header = (
            target_headers[column_offset]
            if column_offset < target_width
            else f"Column{column_offset + 1}"
        )
        output_header.extend([f"{source_header}_ECC", f"{target_header}_S4"])
    output_worksheet.append(output_header)

    matched_count = 0
    not_matched_count = 0

    for row_index in range(pair_count):
        source_row = (
            source_rows_sorted[row_index]
            if row_index < len(source_rows_sorted)
            else tuple(None for _ in range(source_width))
        )
        target_row = (
            target_rows_sorted[row_index]
            if row_index < len(target_rows_sorted)
            else tuple(None for _ in range(target_width))
        )

        source_row_padded = _pad_row(source_row, compare_width)
        target_row_padded = _pad_row(target_row, compare_width)

        is_matched = _normalized_row(source_row_padded) == _normalized_row(target_row_padded)
        status = "Matched" if is_matched else "Not Matched"

        if is_matched:
            matched_count += 1
        else:
            not_matched_count += 1

        output_row = [status]
        for column_offset in range(compare_width):
            source_value = source_row_padded[column_offset] if column_offset < source_width else None
            target_value = target_row_padded[column_offset] if column_offset < target_width else None
            output_row.extend([source_value, target_value])
        output_worksheet.append(output_row)

        excel_row_number = row_index + 2
        pair_checks: list[str] = []
        for column_offset in range(compare_width):
            source_cell_column = 2 + (column_offset * 2)
            target_cell_column = source_cell_column + 1
            source_cell_ref = f"{get_column_letter(source_cell_column)}{excel_row_number}"
            target_cell_ref = f"{get_column_letter(target_cell_column)}{excel_row_number}"
            pair_checks.append(f"{source_cell_ref}={target_cell_ref}")

        if pair_checks:
            status_formula = f'=IF(AND({",".join(pair_checks)}),"Matched","Not Matched")'
        else:
            status_formula = '="Matched"'

        output_worksheet.cell(row=excel_row_number, column=1).value = status_formula

    if pair_count > 0 and compare_width > 0:
        last_output_row = pair_count + 1
        for column_offset in range(compare_width):
            source_cell_column = 2 + (column_offset * 2)
            target_cell_column = source_cell_column + 1

            source_col_letter = get_column_letter(source_cell_column)
            target_col_letter = get_column_letter(target_cell_column)

            source_range = f"{source_col_letter}2:{source_col_letter}{last_output_row}"
            target_range = f"{target_col_letter}2:{target_col_letter}{last_output_row}"
            mismatch_formula = [f"${source_col_letter}2<>${target_col_letter}2"]

            output_worksheet.conditional_formatting.add(
                source_range,
                FormulaRule(formula=mismatch_formula, font=RED_BOLD_FONT),
            )
            output_worksheet.conditional_formatting.add(
                target_range,
                FormulaRule(formula=mismatch_formula, font=RED_BOLD_FONT),
            )

    _style_output_worksheet(output_worksheet)
    output_worksheet.freeze_panes = "A2"
    _auto_size_columns(output_worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(output_path)
    output_workbook.close()

    return {
        "source_file": str(source_path),
        "target_file": str(target_path),
        "output_file": str(output_path),
        "matched_rows": matched_count,
        "not_matched_rows": not_matched_count,
        "processed_rows": pair_count,
    }
